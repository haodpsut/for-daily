"""Dump current DB state → xlsx (filled in import template format).

Mục đích: round-trip test cho import service.
Sinh file xlsx có sẵn data hiện trong DB, theo đúng format mà import_excel.py
mong đợi → có thể import lại để verify schema chuẩn.

Run:
    python scripts/seed_to_xlsx.py --program-code 7480201 \
        --out ../import_templates/CNTT_7480201_filled.xlsx
"""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from app.db import SessionLocal
from app.models import Program, PO, PLO, PI, PLO_PO, Course, CLO, CLO_PI, Assessment, Assessment_CLO, WeeklyPlan, WeeklyPlan_CLO

HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _write_sheet(wb: Workbook, name: str, headers: list[str], rows: list[list]):
    ws = wb.create_sheet(name)
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(1, c, h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(rows, start=2):
        for c, val in enumerate(row, start=1):
            ws.cell(r, c, val)
    for c in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = max(15, len(headers[c - 1]) + 4)
    ws.freeze_panes = "A2"


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--program-code", default="7480201")
    p.add_argument("--out", required=True)
    args = p.parse_args()

    db = SessionLocal()
    try:
        program = db.query(Program).filter_by(code=args.program_code).first()
        if not program:
            print(f"[ERR] Program {args.program_code} not found")
            sys.exit(1)

        wb = Workbook()
        wb.remove(wb.active)

        # 00_Program
        _write_sheet(wb, "00_Program",
            ["code", "name_vn", "name_en", "level", "duration_years",
             "total_credits", "language", "decision_no", "decision_date", "issuing_authority"],
            [[program.code, program.name_vn, program.name_en, program.level.value,
              program.duration_years, program.total_credits, program.language,
              program.decision_no,
              program.decision_date.isoformat() if program.decision_date else None,
              program.issuing_authority]])

        # 01_PO
        pos = sorted(program.pos, key=lambda x: x.order)
        _write_sheet(wb, "01_PO",
            ["code", "text_vn", "text_en", "order"],
            [[po.code, po.text_vn, po.text_en, po.order] for po in pos])

        # 02_PLO
        plos = sorted(program.plos, key=lambda x: x.order)
        _write_sheet(wb, "02_PLO",
            ["code", "text_vn", "text_en", "order"],
            [[plo.code, plo.text_vn, plo.text_en, plo.order] for plo in plos])

        # 03_PI
        pi_rows = []
        for plo in plos:
            for pi in sorted(plo.pis, key=lambda x: x.order):
                pi_rows.append([pi.code, plo.code, pi.text_vn, pi.text_en, pi.order])
        _write_sheet(wb, "03_PI",
            ["code", "plo_code", "text_vn", "text_en", "order"],
            pi_rows)

        # 04_PLO_PO_matrix
        plo_po_rows = (
            db.query(PLO_PO.plo_id, PO.code)
              .join(PO, PLO_PO.po_id == PO.id)
              .all()
        )
        plo_po_map: dict[str, set[str]] = {}
        plo_id_to_code = {p.id: p.code for p in plos}
        for plo_id, po_code in plo_po_rows:
            if plo_id in plo_id_to_code:
                plo_po_map.setdefault(plo_id_to_code[plo_id], set()).add(po_code)

        po_codes = [po.code for po in pos]
        matrix_rows = []
        for plo in plos:
            row = [plo.code]
            for po_code in po_codes:
                row.append("X" if po_code in plo_po_map.get(plo.code, set()) else None)
            matrix_rows.append(row)
        _write_sheet(wb, "04_PLO_PO_matrix",
            ["plo_code"] + po_codes,
            matrix_rows)

        # 05_PLO_VQF_matrix — empty for now
        _write_sheet(wb, "05_PLO_VQF_matrix",
            ["plo_code", "K1", "K2", "K3", "K4", "K5",
             "S1", "S2", "S3", "S4", "S5", "S6",
             "A1", "A2", "A3", "A4"],
            [])

        # 06_Course
        courses = sorted(program.courses, key=lambda c: (c.semester_default or 99, c.code))
        _write_sheet(wb, "06_Course",
            ["code", "name_vn", "name_en", "credits",
             "hours_lt", "hours_th", "hours_self",
             "prerequisites", "corequisites",
             "knowledge_group", "is_elective", "semester_default",
             "language", "description"],
            [[c.code, c.name_vn, c.name_en, c.credits,
              c.hours_lt, c.hours_th, c.hours_self,
              c.prerequisites, c.corequisites,
              c.knowledge_group.value, c.is_elective, c.semester_default,
              c.language, c.description] for c in courses])

        # 07_CLO
        clo_rows = []
        for course in courses:
            for clo in sorted(course.clos, key=lambda x: x.order):
                clo_rows.append([course.code, clo.code, clo.text_vn, clo.text_en, clo.order])
        _write_sheet(wb, "07_CLO",
            ["course_code", "clo_code", "text_vn", "text_en", "order"],
            clo_rows)

        # 08_CLO_PI_matrix
        clo_pi_rows = []
        course_by_id = {c.id: c for c in courses}
        clo_to_course_code = {clo.id: course_by_id[clo.course_id].code
                              for course in courses for clo in course.clos}
        clo_to_clo_code = {clo.id: clo.code for course in courses for clo in course.clos}
        pi_id_to_code = {pi.id: pi.code for plo in plos for pi in plo.pis}
        for cp in db.query(CLO_PI).all():
            if cp.clo_id in clo_to_course_code:
                clo_pi_rows.append([
                    clo_to_course_code[cp.clo_id],
                    clo_to_clo_code[cp.clo_id],
                    pi_id_to_code.get(cp.pi_id),
                    cp.level.value,
                ])
        _write_sheet(wb, "08_CLO_PI_matrix",
            ["course_code", "clo_code", "pi_code", "level"],
            clo_pi_rows)

        # 09_Assessment
        asmt_rows = []
        for course in courses:
            for asmt in sorted(course.assessments, key=lambda x: x.order):
                clo_codes_for_asmt = (
                    db.query(CLO.code)
                      .join(Assessment_CLO, Assessment_CLO.clo_id == CLO.id)
                      .filter(Assessment_CLO.assessment_id == asmt.id)
                      .all()
                )
                clo_codes_str = ",".join(c[0] for c in clo_codes_for_asmt)
                asmt_rows.append([course.code, asmt.component_name, asmt.weight_pct,
                                  asmt.method, clo_codes_str, asmt.order])
        _write_sheet(wb, "09_Assessment",
            ["course_code", "component_name", "weight_pct", "method", "clo_codes", "order"],
            asmt_rows)

        # 10_WeeklyPlan
        wp_rows = []
        for course in courses:
            for wp in sorted(course.weekly_plans, key=lambda x: x.week):
                clo_codes_for_wp = (
                    db.query(CLO.code)
                      .join(WeeklyPlan_CLO, WeeklyPlan_CLO.clo_id == CLO.id)
                      .filter(WeeklyPlan_CLO.weekly_plan_id == wp.id)
                      .all()
                )
                clo_codes_str = ",".join(c[0] for c in clo_codes_for_wp)
                wp_rows.append([course.code, wp.week, wp.topic, wp.content,
                                wp.hours_lt, wp.hours_th, clo_codes_str])
        _write_sheet(wb, "10_WeeklyPlan",
            ["course_code", "week", "topic", "content", "hours_lt", "hours_th", "clo_codes"],
            wp_rows)

        out_path = Path(args.out)
        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        print(f"[OK] Wrote {out_path.resolve()}")
        print(f"     {len(pos)} POs, {len(plos)} PLOs, {len(pi_rows)} PIs, "
              f"{sum(len(r) - 1 for r in matrix_rows)} matrix marks (raw), "
              f"{len(courses)} courses")

    finally:
        db.close()


if __name__ == "__main__":
    main()

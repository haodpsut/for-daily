"""Sinh Excel template trống cho import CTĐT.

Usage:
    python scripts/gen_import_template.py --program-code 7480201 \
        --out ../../import_templates/CNTT_7480201_template.xlsx
"""
import argparse
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
HINT_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")


SHEETS = [
    {
        "name": "00_Program",
        "headers": ["code", "name_vn", "name_en", "level", "duration_years",
                    "total_credits", "language", "decision_no", "decision_date",
                    "issuing_authority"],
        "example": ["7480201", "Công nghệ thông tin", "Information Technology",
                    "DAI_HOC", 4, 144, "Tiếng Việt", "346/QĐ-ĐHKTĐN",
                    "2024-06-25", "Trường ĐH Kiến trúc Đà Nẵng"],
    },
    {
        "name": "01_PO",
        "headers": ["code", "text_vn", "text_en", "order"],
        "example": ["PO1", "Đào tạo nguồn nhân lực có phẩm chất chính trị, đạo đức xã hội...", "", 1],
    },
    {
        "name": "02_PLO",
        "headers": ["code", "text_vn", "text_en", "order"],
        "example": ["PLO1", "Áp dụng các kiến thức cơ bản về khoa học xã hội, kinh tế và pháp luật...", "", 1],
    },
    {
        "name": "03_PI",
        "headers": ["code", "plo_code", "text_vn", "text_en", "order"],
        "example": ["PI1.1", "PLO1", "Vận dụng các kiến thức nền tảng về lý luận chính trị...", "", 1],
    },
    {
        "name": "04_PLO_PO_matrix",
        "headers": ["plo_code", "PO1", "PO2", "PO3", "PO4", "PO5", "PO6"],
        "example": ["PLO1", "X", "", "", "", "", ""],
    },
    {
        "name": "05_PLO_VQF_matrix",
        "headers": ["plo_code",
                    "K1", "K2", "K3", "K4", "K5",
                    "S1", "S2", "S3", "S4", "S5", "S6",
                    "A1", "A2", "A3", "A4"],
        "example": ["PLO1", "X", "", "", "", "", "", "", "", "", "", "", "", "", "", ""],
    },
    {
        "name": "06_Course",
        "headers": ["code", "name_vn", "name_en", "credits",
                    "hours_lt", "hours_th", "hours_self",
                    "prerequisites", "corequisites",
                    "knowledge_group", "is_elective", "semester_default",
                    "language", "description"],
        "example": ["MAT101", "Toán cao cấp 1", "Calculus 1", 3,
                    30, 15, 90, "", "",
                    "DAI_CUONG", False, 1,
                    "Tiếng Việt", "Học phần cung cấp kiến thức về giới hạn..."],
    },
    {
        "name": "07_CLO",
        "headers": ["course_code", "clo_code", "text_vn", "text_en", "order"],
        "example": ["MAT101", "CLO1", "Tính được giới hạn của dãy số và hàm số...", "", 1],
    },
    {
        "name": "08_CLO_PI_matrix",
        "headers": ["course_code", "clo_code", "pi_code", "level"],
        "example": ["MAT101", "CLO1", "PI1.1", "I"],
    },
    {
        "name": "09_Assessment",
        "headers": ["course_code", "component_name", "weight_pct",
                    "method", "clo_codes", "order"],
        "example": ["MAT101", "Thi cuối kỳ", 60, "Tự luận 90p", "CLO1,CLO2,CLO3", 3],
    },
    {
        "name": "10_WeeklyPlan",
        "headers": ["course_code", "week", "topic", "content",
                    "hours_lt", "hours_th", "clo_codes"],
        "example": ["MAT101", 1, "Giới hạn dãy số",
                    "Định nghĩa, các tính chất cơ bản, ví dụ", 3, 0, "CLO1"],
    },
]


README_TEXT = """\
HƯỚNG DẪN ĐIỀN

1. Mở từng sheet (00_Program → 10_WeeklyPlan), điền dữ liệu thực tế của CTĐT.
2. Hàng đầu (header xanh) — KHÔNG SỬA tên cột.
3. Hàng thứ 2 (vàng) — DÒNG VÍ DỤ, có thể xoá hoặc thay bằng dữ liệu thật.
4. Mã (code) phải nhất quán giữa các sheet:
   - PI.plo_code phải khớp PLO.code
   - CLO.course_code phải khớp Course.code
   - CLO_PI_matrix.pi_code phải khớp PI.code
5. Enum hợp lệ:
   - level: DAI_HOC | THAC_SI | TIEN_SI
   - knowledge_group: DAI_CUONG | CO_SO | CHUYEN_NGANH | TU_CHON | TOT_NGHIEP
   - CLO_PI_matrix.level: I | R | M | A
6. Ma trận:
   - 04_PLO_PO_matrix: cell = "X" hoặc trống
   - 05_PLO_VQF_matrix: cell = "X" hoặc trống
7. Tổng weight_pct mỗi course ở sheet 09_Assessment phải = 100.

UPLOAD:
    POST http://localhost:8000/api/import/excel
    Form: file=<this xlsx>, program_code=<code>
"""


def make_workbook() -> Workbook:
    wb = Workbook()
    wb.remove(wb.active)

    readme = wb.create_sheet("README")
    for i, line in enumerate(README_TEXT.splitlines(), start=1):
        readme.cell(row=i, column=1, value=line)
    readme.column_dimensions["A"].width = 100

    for sheet_def in SHEETS:
        ws = wb.create_sheet(sheet_def["name"])
        headers = sheet_def["headers"]
        for col, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        for col, val in enumerate(sheet_def["example"], start=1):
            cell = ws.cell(row=2, column=col, value=val)
            cell.fill = HINT_FILL

        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = max(15, len(headers[col - 1]) + 4)
        ws.freeze_panes = "A2"

    return wb


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--program-code", default="7480201", help="Mã ngành (chỉ dùng cho tên file gợi ý)")
    p.add_argument("--out", required=True, help="Đường dẫn xuất file .xlsx")
    args = p.parse_args()

    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = make_workbook()
    wb.save(out_path)
    print(f"OK: wrote {out_path.resolve()}")
    print(f"     11 data sheets + README, 1 example row each")


if __name__ == "__main__":
    main()

"""Import Excel gradebook → upsert students, questions, scores cho 1 session.

Sheets expected (xem `gen_import_template.py`):
- 'Sinh viên': code | full_name | cohort_code | email | absent
- 'Câu hỏi'  : number | text | max_score | bloom | clo_codes | weight
- 'Điểm'     : student_code | Q<number> | Q<number> | ...
"""
from __future__ import annotations

from datetime import datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy.orm import Session as DbSession

from ..models.measurement import (
    BloomLevel,
    MeasQuestion,
    MeasQuestionCLO,
    MeasScore,
    MeasSession,
    MeasSessionStudent,
    MeasStudent,
)
from ..models import CLO


def _parse_decimal(value, default: Decimal | None = None) -> Decimal | None:
    if value is None or value == "":
        return default
    try:
        return Decimal(str(value).strip())
    except (InvalidOperation, ValueError):
        return default


def _parse_bool(value) -> bool:
    if value is None:
        return False
    s = str(value).strip().upper()
    return s in ("TRUE", "1", "YES", "Y", "T", "X", "Đ")


def import_gradebook(
    db: DbSession,
    session: MeasSession,
    file_bytes: bytes,
    user_id: str,
) -> dict:
    """Import 1 file Excel vào session đã tồn tại.

    Idempotent: students upsert by (program_id, code); scores upsert by (student, question);
    questions upsert by (session, number) — nhưng không xoá questions cũ nếu sheet thiếu.
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)

    summary = {
        "students_created": 0,
        "students_updated": 0,
        "students_enrolled": 0,
        "questions_created": 0,
        "questions_updated": 0,
        "scores_created": 0,
        "scores_updated": 0,
        "warnings": [],
    }
    warnings: list[str] = summary["warnings"]

    program_id = session.program_id
    course_id = session.course_id

    # Pre-load existing students of this program by code
    existing_students = {
        s.code: s
        for s in db.query(MeasStudent).filter_by(program_id=program_id).all()
    }
    # Pre-load CLOs of this course by code
    course_clos = {
        c.code: c for c in db.query(CLO).filter_by(course_id=course_id).all()
    }

    # ------------------------------------------------------------------
    # Sheet 1: Sinh viên
    # ------------------------------------------------------------------
    if "Sinh viên" not in wb.sheetnames:
        warnings.append("Thiếu sheet 'Sinh viên'")
    else:
        ws = wb["Sinh viên"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            warnings.append("Sheet 'Sinh viên' trống")
        else:
            header = [str(c).strip() if c else "" for c in rows[0]]
            try:
                col_code = header.index("code")
                col_name = header.index("full_name")
            except ValueError as e:
                warnings.append(f"Sheet 'Sinh viên' thiếu cột bắt buộc: {e}")
                col_code = col_name = -1

            col_cohort = header.index("cohort_code") if "cohort_code" in header else -1
            col_email = header.index("email") if "email" in header else -1
            col_absent = header.index("absent") if "absent" in header else -1

            if col_code >= 0:
                enrolled_ids = {link.student_id for link in session.students}
                for row in rows[1:]:
                    if not row or not row[col_code]:
                        continue
                    code = str(row[col_code]).strip()
                    if not code:
                        continue
                    name = str(row[col_name] or "").strip() if col_name >= 0 else ""
                    cohort = str(row[col_cohort] or "").strip() if col_cohort >= 0 else None
                    email = str(row[col_email] or "").strip() if col_email >= 0 else None
                    absent = _parse_bool(row[col_absent]) if col_absent >= 0 else False

                    if code in existing_students:
                        st = existing_students[code]
                        if name and st.full_name != name:
                            st.full_name = name
                        if cohort:
                            st.cohort_code = cohort
                        if email:
                            st.email = email
                        summary["students_updated"] += 1
                    else:
                        st = MeasStudent(
                            program_id=program_id,
                            code=code,
                            full_name=name or code,
                            cohort_code=cohort,
                            email=email,
                        )
                        db.add(st)
                        db.flush()
                        existing_students[code] = st
                        summary["students_created"] += 1

                    # Enroll vào session nếu chưa
                    if st.id not in enrolled_ids:
                        db.add(
                            MeasSessionStudent(
                                session_id=session.id, student_id=st.id, absent=absent
                            )
                        )
                        enrolled_ids.add(st.id)
                        summary["students_enrolled"] += 1
                    else:
                        # Update absent flag
                        link = (
                            db.query(MeasSessionStudent)
                            .filter_by(session_id=session.id, student_id=st.id)
                            .first()
                        )
                        if link and link.absent != absent:
                            link.absent = absent
                db.flush()

    # ------------------------------------------------------------------
    # Sheet 2: Câu hỏi
    # ------------------------------------------------------------------
    questions_by_number: dict[str, MeasQuestion] = {q.number: q for q in session.questions}
    if "Câu hỏi" in wb.sheetnames:
        ws = wb["Câu hỏi"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) >= 2:
            header = [str(c).strip() if c else "" for c in rows[0]]
            try:
                col_num = header.index("number")
                col_max = header.index("max_score")
            except ValueError as e:
                warnings.append(f"Sheet 'Câu hỏi' thiếu cột bắt buộc: {e}")
                col_num = col_max = -1

            col_text = header.index("text") if "text" in header else -1
            col_bloom = header.index("bloom") if "bloom" in header else -1
            col_clos = header.index("clo_codes") if "clo_codes" in header else -1
            col_weight = header.index("weight") if "weight" in header else -1

            if col_num >= 0:
                order = 0
                for row in rows[1:]:
                    if not row or row[col_num] is None or row[col_num] == "":
                        continue
                    number = str(row[col_num]).strip()
                    max_s = _parse_decimal(row[col_max], Decimal("0"))
                    if max_s is None or max_s <= 0:
                        warnings.append(f"Câu '{number}': max_score không hợp lệ")
                        continue
                    text = str(row[col_text] or "").strip() if col_text >= 0 else None
                    bloom_str = str(row[col_bloom] or "").strip().upper() if col_bloom >= 0 else ""
                    bloom = None
                    if bloom_str and bloom_str in BloomLevel.__members__:
                        bloom = BloomLevel[bloom_str]
                    weight = (
                        _parse_decimal(row[col_weight]) if col_weight >= 0 else None
                    )

                    order += 1
                    if number in questions_by_number:
                        q = questions_by_number[number]
                        q.text = text or q.text
                        q.max_score = max_s
                        q.bloom_level = bloom or q.bloom_level
                        q.weight_in_session = weight
                        summary["questions_updated"] += 1
                    else:
                        q = MeasQuestion(
                            session_id=session.id,
                            number=number,
                            order=order,
                            text=text,
                            max_score=max_s,
                            bloom_level=bloom,
                            weight_in_session=weight,
                        )
                        db.add(q)
                        db.flush()
                        questions_by_number[number] = q
                        summary["questions_created"] += 1

                    # CLO links: format "CLO1:0.7, CLO2:0.3" or "CLO1, CLO2"
                    if col_clos >= 0 and row[col_clos]:
                        parts = [p.strip() for p in str(row[col_clos]).split(",") if p.strip()]
                        # Replace existing links
                        db.query(MeasQuestionCLO).filter_by(question_id=q.id).delete()
                        for part in parts:
                            if ":" in part:
                                code, w = part.split(":", 1)
                                code = code.strip()
                                w_dec = _parse_decimal(w, Decimal("1.0")) or Decimal("1.0")
                            else:
                                code, w_dec = part, Decimal("1.0")
                            clo = course_clos.get(code)
                            if not clo:
                                warnings.append(
                                    f"Câu '{number}': CLO code '{code}' không thuộc course"
                                )
                                continue
                            db.add(
                                MeasQuestionCLO(
                                    question_id=q.id, clo_id=clo.id, weight=w_dec
                                )
                            )
                db.flush()

    # ------------------------------------------------------------------
    # Sheet 3: Điểm — matrix student × question
    # ------------------------------------------------------------------
    if "Điểm" not in wb.sheetnames:
        warnings.append("Thiếu sheet 'Điểm'")
    else:
        ws = wb["Điểm"]
        rows = list(ws.iter_rows(values_only=True))
        if len(rows) < 2:
            warnings.append("Sheet 'Điểm' trống")
        else:
            header = [str(c).strip() if c else "" for c in rows[0]]
            if not header or header[0] != "student_code":
                warnings.append("Sheet 'Điểm': cột đầu phải là 'student_code'")
            else:
                # Map column index → question
                col_to_question: dict[int, MeasQuestion] = {}
                for idx in range(1, len(header)):
                    col_label = header[idx]
                    if not col_label:
                        continue
                    # accept "Q1" or just "1"
                    qnum = col_label[1:] if col_label.upper().startswith("Q") else col_label
                    qnum = qnum.strip()
                    if qnum in questions_by_number:
                        col_to_question[idx] = questions_by_number[qnum]
                    else:
                        warnings.append(
                            f"Sheet 'Điểm': cột '{col_label}' không khớp question nào"
                        )

                # Pre-load existing scores
                existing_scores = {
                    (s.student_id, s.question_id): s
                    for s in db.query(MeasScore).filter_by(session_id=session.id).all()
                }

                for row in rows[1:]:
                    if not row or not row[0]:
                        continue
                    code = str(row[0]).strip()
                    student = existing_students.get(code)
                    if not student:
                        warnings.append(f"Sheet 'Điểm': student_code '{code}' chưa tồn tại")
                        continue
                    for idx, q in col_to_question.items():
                        if idx >= len(row):
                            continue
                        value = row[idx]
                        if value is None or value == "":
                            continue  # bỏ qua chưa chấm
                        raw = _parse_decimal(value)
                        if raw is None:
                            warnings.append(
                                f"SV {code}, câu {q.number}: điểm '{value}' không hợp lệ"
                            )
                            continue
                        if raw < 0 or raw > q.max_score:
                            warnings.append(
                                f"SV {code}, câu {q.number}: điểm {raw} ngoài [0, {q.max_score}]"
                            )
                            continue
                        existing = existing_scores.get((student.id, q.id))
                        if existing:
                            if existing.raw_score != raw:
                                existing.raw_score = raw
                                existing.graded_at = datetime.utcnow()
                                existing.graded_by = user_id
                                summary["scores_updated"] += 1
                        else:
                            db.add(
                                MeasScore(
                                    session_id=session.id,
                                    student_id=student.id,
                                    question_id=q.id,
                                    raw_score=raw,
                                    graded_at=datetime.utcnow(),
                                    graded_by=user_id,
                                )
                            )
                            summary["scores_created"] += 1

    db.commit()
    # Cap warnings
    if len(warnings) > 30:
        summary["warnings"] = warnings[:30] + [f"... ({len(warnings) - 30} more)"]
    return summary

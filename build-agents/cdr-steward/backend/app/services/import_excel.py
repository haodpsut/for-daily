"""Excel → DB import service.

Đọc file .xlsx (sinh bởi gen_import_template.py), validate cross-references
giữa các sheet, ghi vào DB trong 1 transaction (rollback nếu lỗi).

Behavior:
- Nếu Program với cùng `code` đã tồn tại → DELETE rồi tạo mới (full upsert).
  → Tránh drift: 1 chương trình = 1 nguồn dữ liệu duy nhất.
- Sheet trống → bỏ qua, không lỗi.
- Validation lỗi → KHÔNG ghi gì, trả về danh sách lỗi.
- Validation cảnh báo → vẫn ghi, kèm warnings.
"""
from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from ..models import (
    Program, PO, PLO, PI, PLO_PO,
    Course, CLO, CLO_PI,
    Assessment, Assessment_CLO,
    WeeklyPlan, WeeklyPlan_CLO,
    ProgramLevel, KnowledgeGroup, IRMALevel,
)


# ─────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────

def _read_sheet(ws) -> list[dict[str, Any]]:
    """Đọc sheet → list[dict], dùng row 1 làm header, bỏ qua row trống."""
    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    headers = [str(h).strip() if h is not None else "" for h in headers]
    rows = []
    for r in range(2, ws.max_row + 1):
        values = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in values):
            continue
        rows.append(dict(zip(headers, values)))
    return rows


def _str(v) -> str | None:
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def _int(v, default=None) -> int | None:
    if v is None or (isinstance(v, str) and not v.strip()):
        return default
    return int(v)


def _bool(v) -> bool:
    if isinstance(v, bool):
        return v
    if v is None:
        return False
    return str(v).strip().lower() in ("true", "1", "yes", "x", "có")


def _parse_date(v) -> date | None:
    if v is None or v == "":
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    if isinstance(v, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
            try:
                return datetime.strptime(v.strip(), fmt).date()
            except ValueError:
                continue
    raise ValueError(f"Cannot parse date: {v!r}")


def _is_x(v) -> bool:
    return v is not None and str(v).strip().upper() == "X"


def _split_codes(v) -> list[str]:
    if v is None:
        return []
    return [c.strip() for c in str(v).split(",") if c.strip()]


# ─────────────────────────────────────────────────────────────
# Main import
# ─────────────────────────────────────────────────────────────

def import_excel(db: Session, xlsx_path: Path) -> dict:
    wb = load_workbook(xlsx_path, data_only=True)

    # 1) Read all sheets (skip README)
    sheets: dict[str, list[dict]] = {}
    expected = [
        "00_Program", "01_PO", "02_PLO", "03_PI",
        "04_PLO_PO_matrix", "05_PLO_VQF_matrix",
        "06_Course", "07_CLO", "08_CLO_PI_matrix",
        "09_Assessment", "10_WeeklyPlan",
    ]
    for name in expected:
        if name in wb.sheetnames:
            sheets[name] = _read_sheet(wb[name])
        else:
            sheets[name] = []

    # 2) Validate
    errors: list[str] = []
    warnings: list[str] = []

    if not sheets["00_Program"]:
        return {"imported": {}, "warnings": [], "errors": ["Sheet 00_Program rỗng"]}
    if len(sheets["00_Program"]) > 1:
        errors.append("Sheet 00_Program chỉ được có 1 dòng dữ liệu")

    program_row = sheets["00_Program"][0]
    if not program_row.get("code"):
        errors.append("Program: thiếu 'code'")

    plo_codes = {r["code"] for r in sheets["02_PLO"] if r.get("code")}
    po_codes = {r["code"] for r in sheets["01_PO"] if r.get("code")}
    pi_codes = {r["code"] for r in sheets["03_PI"] if r.get("code")}
    course_codes = {r["code"] for r in sheets["06_Course"] if r.get("code")}
    clo_pairs = {(r.get("course_code"), r.get("clo_code")) for r in sheets["07_CLO"]}

    for r in sheets["03_PI"]:
        if r.get("plo_code") not in plo_codes:
            errors.append(f"PI {r.get('code')}: plo_code '{r.get('plo_code')}' không có trong sheet 02_PLO")

    for r in sheets["04_PLO_PO_matrix"]:
        if r.get("plo_code") not in plo_codes:
            errors.append(f"PLO_PO_matrix: plo_code '{r.get('plo_code')}' không có trong sheet 02_PLO")
        for po_code in po_codes:
            v = r.get(po_code)
            if v is not None and str(v).strip() and not _is_x(v):
                errors.append(f"PLO_PO_matrix [{r.get('plo_code')},{po_code}]: chỉ chấp nhận 'X' hoặc trống, gặp '{v}'")

    for r in sheets["07_CLO"]:
        if r.get("course_code") not in course_codes:
            errors.append(f"CLO {r.get('clo_code')}: course_code '{r.get('course_code')}' không có trong sheet 06_Course")

    for r in sheets["08_CLO_PI_matrix"]:
        pair = (r.get("course_code"), r.get("clo_code"))
        if pair not in clo_pairs:
            errors.append(f"CLO_PI_matrix: cặp ({pair[0]}, {pair[1]}) không có trong sheet 07_CLO")
        if r.get("pi_code") not in pi_codes:
            errors.append(f"CLO_PI_matrix: pi_code '{r.get('pi_code')}' không có trong sheet 03_PI")
        lvl = _str(r.get("level"))
        if lvl not in ("I", "R", "M", "A"):
            errors.append(f"CLO_PI_matrix: level '{lvl}' không hợp lệ (I/R/M/A)")

    course_weights: dict[str, int] = {}
    for r in sheets["09_Assessment"]:
        cc = r.get("course_code")
        if cc:
            course_weights[cc] = course_weights.get(cc, 0) + (_int(r.get("weight_pct"), 0) or 0)
    for cc, total in course_weights.items():
        if total != 100:
            warnings.append(f"Course {cc}: tổng weight_pct = {total} (mong đợi 100)")

    if errors:
        return {"imported": {}, "warnings": warnings, "errors": errors}

    # 3) Write to DB (transactional)
    try:
        existing = db.query(Program).filter_by(code=program_row["code"]).first()
        if existing:
            db.delete(existing)
            db.flush()

        program = Program(
            code=str(program_row["code"]).strip(),
            name_vn=_str(program_row.get("name_vn")) or "",
            name_en=_str(program_row.get("name_en")),
            level=ProgramLevel(_str(program_row.get("level")) or "DAI_HOC"),
            duration_years=_int(program_row.get("duration_years"), 4) or 4,
            total_credits=_int(program_row.get("total_credits")),
            language=_str(program_row.get("language")) or "Tiếng Việt",
            decision_no=_str(program_row.get("decision_no")),
            decision_date=_parse_date(program_row.get("decision_date")),
            issuing_authority=_str(program_row.get("issuing_authority")),
        )
        db.add(program)
        db.flush()

        po_by_code: dict[str, PO] = {}
        for r in sheets["01_PO"]:
            po = PO(
                program_id=program.id,
                code=r["code"],
                text_vn=_str(r.get("text_vn")) or "",
                text_en=_str(r.get("text_en")),
                order=_int(r.get("order"), 0) or 0,
            )
            db.add(po)
            po_by_code[r["code"]] = po
        db.flush()

        plo_by_code: dict[str, PLO] = {}
        for r in sheets["02_PLO"]:
            plo = PLO(
                program_id=program.id,
                code=r["code"],
                text_vn=_str(r.get("text_vn")) or "",
                text_en=_str(r.get("text_en")),
                order=_int(r.get("order"), 0) or 0,
            )
            db.add(plo)
            plo_by_code[r["code"]] = plo
        db.flush()

        pi_by_code: dict[str, PI] = {}
        for r in sheets["03_PI"]:
            pi = PI(
                plo_id=plo_by_code[r["plo_code"]].id,
                code=r["code"],
                text_vn=_str(r.get("text_vn")) or "",
                text_en=_str(r.get("text_en")),
                order=_int(r.get("order"), 0) or 0,
            )
            db.add(pi)
            pi_by_code[r["code"]] = pi
        db.flush()

        plo_po_count = 0
        for r in sheets["04_PLO_PO_matrix"]:
            for po_code, po_obj in po_by_code.items():
                if _is_x(r.get(po_code)):
                    db.add(PLO_PO(
                        plo_id=plo_by_code[r["plo_code"]].id,
                        po_id=po_obj.id,
                    ))
                    plo_po_count += 1

        course_by_code: dict[str, Course] = {}
        for r in sheets["06_Course"]:
            course = Course(
                program_id=program.id,
                code=r["code"],
                name_vn=_str(r.get("name_vn")) or "",
                name_en=_str(r.get("name_en")),
                credits=_int(r.get("credits"), 0) or 0,
                hours_lt=_int(r.get("hours_lt"), 0) or 0,
                hours_th=_int(r.get("hours_th"), 0) or 0,
                hours_self=_int(r.get("hours_self"), 0) or 0,
                prerequisites=_str(r.get("prerequisites")),
                corequisites=_str(r.get("corequisites")),
                knowledge_group=KnowledgeGroup(_str(r.get("knowledge_group")) or "CHUYEN_NGANH"),
                is_elective=_bool(r.get("is_elective")),
                semester_default=_int(r.get("semester_default")),
                language=_str(r.get("language")) or "Tiếng Việt",
                description=_str(r.get("description")),
            )
            db.add(course)
            course_by_code[r["code"]] = course
        db.flush()

        clo_by_pair: dict[tuple[str, str], CLO] = {}
        for r in sheets["07_CLO"]:
            clo = CLO(
                course_id=course_by_code[r["course_code"]].id,
                code=r["clo_code"],
                text_vn=_str(r.get("text_vn")) or "",
                text_en=_str(r.get("text_en")),
                order=_int(r.get("order"), 0) or 0,
            )
            db.add(clo)
            clo_by_pair[(r["course_code"], r["clo_code"])] = clo
        db.flush()

        for r in sheets["08_CLO_PI_matrix"]:
            db.add(CLO_PI(
                clo_id=clo_by_pair[(r["course_code"], r["clo_code"])].id,
                pi_id=pi_by_code[r["pi_code"]].id,
                level=IRMALevel(_str(r["level"])),
            ))

        for r in sheets["09_Assessment"]:
            asmt = Assessment(
                course_id=course_by_code[r["course_code"]].id,
                component_name=_str(r.get("component_name")) or "",
                weight_pct=_int(r.get("weight_pct"), 0) or 0,
                method=_str(r.get("method")),
                order=_int(r.get("order"), 0) or 0,
            )
            db.add(asmt)
            db.flush()
            for clo_code in _split_codes(r.get("clo_codes")):
                pair = (r["course_code"], clo_code)
                if pair in clo_by_pair:
                    db.add(Assessment_CLO(
                        assessment_id=asmt.id,
                        clo_id=clo_by_pair[pair].id,
                    ))

        for r in sheets["10_WeeklyPlan"]:
            wp = WeeklyPlan(
                course_id=course_by_code[r["course_code"]].id,
                week=_int(r.get("week"), 0) or 0,
                topic=_str(r.get("topic")),
                content=_str(r.get("content")),
                hours_lt=_int(r.get("hours_lt"), 0) or 0,
                hours_th=_int(r.get("hours_th"), 0) or 0,
            )
            db.add(wp)
            db.flush()
            for clo_code in _split_codes(r.get("clo_codes")):
                pair = (r["course_code"], clo_code)
                if pair in clo_by_pair:
                    db.add(WeeklyPlan_CLO(
                        weekly_plan_id=wp.id,
                        clo_id=clo_by_pair[pair].id,
                    ))

        db.commit()

    except Exception as e:
        db.rollback()
        return {"imported": {}, "warnings": warnings, "errors": [f"DB write failed: {e}"]}

    return {
        "imported": {
            "program": 1,
            "po": len(po_by_code),
            "plo": len(plo_by_code),
            "pi": len(pi_by_code),
            "plo_po_mappings": plo_po_count,
            "course": len(course_by_code),
            "clo": len(clo_by_pair),
            "clo_pi_mappings": len(sheets["08_CLO_PI_matrix"]),
            "assessment": len(sheets["09_Assessment"]),
            "weekly_plan": len(sheets["10_WeeklyPlan"]),
        },
        "warnings": warnings,
        "errors": [],
    }

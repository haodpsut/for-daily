"""Sinh Excel gradebook template cho 1 session.

Format: 1 file Excel với 3 sheets
  - Sheet 'Sinh viên' : code | full_name | cohort_code | email | absent (bool)
  - Sheet 'Câu hỏi'   : number | text | max_score | bloom | clo_codes (comma-separated) | weight
  - Sheet 'Điểm'      : student_code | Q1 | Q2 | ... | Qn  (1 row per student, scores nguyên cell)

Usage:
  python scripts/gen_import_template.py --session-id <uuid> --out file.xlsx
  python scripts/gen_import_template.py --blank --questions 5 --students 30 --out blank.xlsx
"""
from __future__ import annotations

import argparse
import os
import sys
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE.parent))

os.environ.setdefault(
    "DATABASE_URL",
    f"sqlite:///{HERE.parent.parent.parent / 'cdr-steward' / 'backend' / 'cdr_steward.db'}",
)

from openpyxl import Workbook  # noqa: E402
from openpyxl.styles import Alignment, Font, PatternFill  # noqa: E402

from app.db import SessionLocal  # noqa: E402
from app.models.meas import MeasSession  # noqa: E402
from app.models.ref import CLO  # noqa: E402

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(bold=True, color="FFFFFF")
NOTE_FILL = PatternFill("solid", fgColor="FFF2CC")


def _style_header(ws, n_cols: int) -> None:
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _style_note(ws, row: int, n_cols: int, text: str) -> None:
    ws.cell(row=row, column=1, value=text).fill = NOTE_FILL
    for col in range(1, n_cols + 1):
        ws.cell(row=row, column=col).fill = NOTE_FILL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=n_cols)


def make_blank(n_questions: int, n_students: int, out_path: Path) -> None:
    wb = Workbook()

    # Sheet 1: Sinh viên
    ws_s = wb.active
    ws_s.title = "Sinh viên"
    headers_s = ["code", "full_name", "cohort_code", "email", "absent"]
    ws_s.append(headers_s)
    _style_header(ws_s, len(headers_s))
    for i in range(1, n_students + 1):
        ws_s.append([f"21QT01{i:03d}", f"Nguyễn Văn {chr(64 + i)}", "21QT01", "", "FALSE"])
    for col, w in zip("ABCDE", [16, 28, 14, 28, 10]):
        ws_s.column_dimensions[col].width = w

    # Sheet 2: Câu hỏi
    ws_q = wb.create_sheet("Câu hỏi")
    headers_q = ["number", "text", "max_score", "bloom", "clo_codes", "weight"]
    ws_q.append(headers_q)
    _style_header(ws_q, len(headers_q))
    bloom_cycle = ["REMEMBER", "UNDERSTAND", "APPLY", "ANALYZE", "EVALUATE"]
    for i in range(1, n_questions + 1):
        ws_q.append(
            [
                str(i),
                f"Đề câu {i} (mô tả ngắn)",
                round(10.0 / n_questions, 2),
                bloom_cycle[(i - 1) % len(bloom_cycle)],
                "CLO1",
                "",
            ]
        )
    for col, w in zip("ABCDEF", [10, 50, 12, 14, 20, 10]):
        ws_q.column_dimensions[col].width = w

    # Sheet 3: Điểm
    ws_g = wb.create_sheet("Điểm")
    headers_g = ["student_code"] + [f"Q{i}" for i in range(1, n_questions + 1)]
    ws_g.append(headers_g)
    _style_header(ws_g, len(headers_g))
    for i in range(1, n_students + 1):
        ws_g.append([f"21QT01{i:03d}"] + [""] * n_questions)
    ws_g.column_dimensions["A"].width = 16

    # Notes
    ws_g.cell(row=n_students + 3, column=1, value="HƯỚNG DẪN:").font = Font(bold=True, italic=True)
    ws_g.cell(
        row=n_students + 4, column=1,
        value="• student_code phải khớp với sheet 'Sinh viên'.",
    )
    ws_g.cell(
        row=n_students + 5, column=1,
        value="• Q1, Q2, ... khớp với 'number' trong sheet 'Câu hỏi'.",
    )
    ws_g.cell(
        row=n_students + 6, column=1,
        value="• Cell trống = chưa chấm; số 0 = đã chấm 0 điểm.",
    )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"✓ Blank template -> {out_path}")


def make_for_session(session_id: str, out_path: Path) -> None:
    db = SessionLocal()
    try:
        sess = db.query(MeasSession).filter_by(id=session_id).first()
        if not sess:
            print(f"✗ Session {session_id} not found")
            sys.exit(1)

        wb = Workbook()

        # Sheet 1: Sinh viên (đã enroll)
        ws_s = wb.active
        ws_s.title = "Sinh viên"
        headers_s = ["code", "full_name", "cohort_code", "email", "absent"]
        ws_s.append(headers_s)
        _style_header(ws_s, len(headers_s))
        students_list = sorted(
            [(link.student, link.absent) for link in sess.students],
            key=lambda t: t[0].code,
        )
        for st, absent in students_list:
            ws_s.append(
                [st.code, st.full_name, st.cohort_code or "", st.email or "", "TRUE" if absent else "FALSE"]
            )
        for col, w in zip("ABCDE", [16, 28, 14, 28, 10]):
            ws_s.column_dimensions[col].width = w

        # Sheet 2: Câu hỏi (đã có)
        ws_q = wb.create_sheet("Câu hỏi")
        headers_q = ["number", "text", "max_score", "bloom", "clo_codes", "weight"]
        ws_q.append(headers_q)
        _style_header(ws_q, len(headers_q))
        questions_sorted = sorted(sess.questions, key=lambda q: q.order)
        for q in questions_sorted:
            clo_codes = []
            for link in q.clo_links:
                clo = db.query(CLO).filter_by(id=link.clo_id).first()
                if clo:
                    clo_codes.append(f"{clo.code}:{link.weight}")
            ws_q.append(
                [
                    q.number,
                    q.text or "",
                    float(q.max_score),
                    q.bloom_level.value if q.bloom_level else "",
                    ", ".join(clo_codes),
                    float(q.weight_in_session) if q.weight_in_session else "",
                ]
            )
        for col, w in zip("ABCDEF", [10, 50, 12, 14, 24, 10]):
            ws_q.column_dimensions[col].width = w

        # Sheet 3: Điểm (matrix điểm sẵn nếu có)
        ws_g = wb.create_sheet("Điểm")
        headers_g = ["student_code"] + [f"Q{q.number}" for q in questions_sorted]
        ws_g.append(headers_g)
        _style_header(ws_g, len(headers_g))

        # Build score matrix
        score_map: dict[tuple[str, str], float] = {}
        for sc in sess.scores:
            if sc.raw_score is not None:
                score_map[(sc.student_id, sc.question_id)] = float(sc.raw_score)
        for st, absent in students_list:
            row = [st.code]
            for q in questions_sorted:
                row.append(score_map.get((st.id, q.id), ""))
            ws_g.append(row)
        ws_g.column_dimensions["A"].width = 16

        out_path.parent.mkdir(parents=True, exist_ok=True)
        wb.save(out_path)
        print(f"✓ Template for session '{sess.name}' -> {out_path}")
        print(f"  {len(students_list)} sinh viên × {len(questions_sorted)} câu")
    finally:
        db.close()


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--session-id", help="Session ID có sẵn để export")
    ap.add_argument("--blank", action="store_true", help="Sinh blank template")
    ap.add_argument("--questions", type=int, default=5, help="Số câu hỏi (blank only)")
    ap.add_argument("--students", type=int, default=30, help="Số SV (blank only)")
    ap.add_argument("--out", required=True, help="File path .xlsx")
    args = ap.parse_args()

    out_path = Path(args.out)
    if args.blank or not args.session_id:
        make_blank(args.questions, args.students, out_path)
    else:
        make_for_session(args.session_id, out_path)


if __name__ == "__main__":
    main()
